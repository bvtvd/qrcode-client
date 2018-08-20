#!/usr/bin/python3
# -*- coding: utf-8 -*-
'''
@Author  :   Hans
@Contact :   hans01@foxmail.com
@File    :   BatchGenerateThread.py
@Time    :   2018/8/7 22:06
@Desc    :   批量生成二维码线程类
'''

from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook
from utils.QRCode import QRCode
import os
import sys

class BatchGenerateThread(QThread):
    signal = pyqtSignal(dict)

    """
    batchFile: 文件路径
    savePath: 生成二维码保存路径
    batchLogoPath: logo 路径
    batchStyle: 样式
    batchSize: 图片大小
    """
    def __init__(self, parent = None):
        super(BatchGenerateThread, self).__init__()
        self.batchFile = parent.batchFile
        self.savePath = parent.batchSavePath
        self.batchLogoPath = parent.batchLogoPath
        self.batchStyle = parent.batchStyle
        self.batchSize = parent.batchSize

    def run(self):
        # 读取文件
        excel = load_workbook(self.batchFile)
        sheetnames = excel.sheetnames
        sheet = excel[sheetnames[0]]
        # 整合表数据
        # 数据格式 [ ['内容', '名字'], [], [] ]
        data = []
        for i in range(2, sheet.max_row + 1):
            item = []
            for j in range(1, sheet.max_column + 1):
                item.append(sheet.cell(row=i, column=j).value)
            data.append(item)

        QRTool = QRCode()

        # 准备进度条数据
        dataLength = len(data)

        success = 0
        fail = 0
        # 生成二维码
        for i, vo in enumerate(data):
            progressBarValue = int(i + 1) / dataLength * 100
            try:
                img = QRTool.make(str(vo[0]), self.batchLogoPath, self.batchStyle)
                imgName = str(i + 1) + '.png'
                try:
                    if vo[1]: imgName = str(vo[1]) + '.png'
                except IndexError:
                    pass

                img.resize((self.batchSize, self.batchSize)).save(os.path.join(self.savePath, imgName))
                message = '<p style="margin:2">{} 生成成功</p>'.format(vo[0])
                success = success + 1
            except:
                print('error: ', sys.exc_info()[0])
                fail = fail + 1
                message = '<p style="margin:2;color:red">{} 生成失败</p>'.format(vo[0])
            self.signal.emit({'progressBarValue': progressBarValue, 'message': message})

        message = '执行完毕, 共处理 {} 条数据, 成功生成 {} 个二维码, 失败 {} 个.'.format(dataLength, success, fail)
        self.signal.emit({'progressBarValue': 100, 'message': message})




